"""Views for browsing and exporting payment requests.

Everything here reads. The router in payments/routers.py raises on any write
attempt, and the `robot` database account holds GRANT SELECT only.
"""

import json
import logging
from datetime import datetime
from functools import wraps

from django.conf import settings
from django.db.models import Count, Q, Sum
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.utils import Error as DatabaseError
from django.http import HttpResponse
from django.shortcuts import render
from django.views.generic import DetailView
from django_filters.views import FilterView

from .filters import PaymentRequestFilter
from .models import PaymentRequest

audit = logging.getLogger("payments.audit")

SETTLED_STATUSES = ["COMPLETED", "PAID"]

LIST_COLUMNS = [
    ("created_at", "Created", True),
    ("payment_reference", "Reference", True),
    ("requester_account_name", "Requester", True),
    ("payer_account_name", "Payer", True),
    ("request_amount", "Amount", True),
    ("status", "Status", True),
    ("request_type", "Type", True),
]

# Rows per page the user may choose between.
PAGE_SIZES = [25, 50, 100, 200]
DEFAULT_PAGE_SIZE = 50

# Query parameters that are not filters, and so never appear as filter chips.
NON_FILTER_PARAMS = {"page", "sort", "per_page"}

EXPORT_COLUMNS = [
    "id", "created_at", "updated_at",
    "payment_reference", "settlement_reference", "cross_bank_token",
    "status", "request_type", "payment_type",
    "request_amount", "request_currency", "description",
    "requester_account_number", "requester_account_name",
    "requester_email", "requester_phone", "requester_customer_id",
    "payer_account_number", "payer_account_name", "payer_display_identifier",
    "payer_email", "payer_phone", "payer_customer_id",
    "cancellation_reason", "decline_reason",
    "created_by_user", "modified_by_user",
    "confirmation_deadline", "reversal_reference",
]


COLUMN_WIDTHS = {
    "id": 38, "cross_bank_token": 38, "settlement_reference": 44,
    "payment_reference": 28, "reversal_reference": 24,
    "created_at": 20, "updated_at": 20, "confirmation_deadline": 20,
    "status": 23, "request_type": 18, "payment_type": 16,
    "request_amount": 16, "request_currency": 9,
    "description": 40,
    "requester_account_number": 16, "payer_account_number": 16,
    "requester_account_name": 32, "payer_account_name": 32,
    "payer_display_identifier": 30,
    "requester_email": 30, "payer_email": 30,
    "requester_phone": 16, "payer_phone": 16,
    "requester_customer_id": 18, "payer_customer_id": 18,
    "cancellation_reason": 30, "decline_reason": 50,
    "created_by_user": 16, "modified_by_user": 16,
}
DEFAULT_WIDTH = 18

DATE_COLUMNS = {"created_at", "updated_at", "confirmation_deadline"}
AMOUNT_COLUMNS = {"request_amount"}
DATE_FORMAT = "yyyy-mm-dd hh:mm:ss"
AMOUNT_FORMAT = "#,##0.00"

PII_COLUMNS = {
    "requester_email", "requester_phone", "requester_customer_id",
    "payer_email", "payer_phone", "payer_customer_id",
}


def handle_db_errors(view):
    """Render the connectivity page instead of a 500 when the CBA is down."""

    @wraps(view)
    def wrapper(request, *args, **kwargs):
        try:
            return view(request, *args, **kwargs)
        except DatabaseError as exc:
            return render(
                request,
                "payments/db_error.html",
                {"error": exc, "host": settings.DATABASES["cba"]["HOST"]},
                status=503,
            )

    return wrapper

DETAIL_GROUPS = [
    ("Request", ["id", "payment_reference", "request_type", "payment_type", "description"]),
    ("Requester", ["requester_account_name", "requester_account_number",
                   "requester_email", "requester_phone", "requester_customer_id"]),
    ("Payer", ["payer_account_name", "payer_account_number", "payer_display_identifier",
               "payer_email", "payer_phone", "payer_customer_id"]),
    ("References", ["cross_bank_token", "settlement_reference", "reversal_reference"]),
    ("Outcome", ["status", "cancellation_reason", "decline_reason"]),
    ("Timeline", ["created_at", "updated_at", "created_by_user", "modified_by_user",
                  "confirmation_deadline", "next_confirmation_reminder_at",
                  "next_poll_at", "next_reversal_at"]),
    ("System", ["version", "poll_attempts", "reversal_attempts",
                "confirmation_reminder_count", "metadata", "payment_details"]),
]

HEADER_FIELDS = {"request_amount", "request_currency"}


class DatabaseErrorMixin:
    """Render a readable page when the core banking database is unreachable.

    Without this a dropped VPN or closed tunnel produces a raw 500, which
    reads like an application bug rather than a connectivity problem.
    """

    def get(self, request, *args, **kwargs):
        try:
            response = super().get(request, *args, **kwargs)
            if hasattr(response, "render"):
                response.render()
            return response
        except DatabaseError as exc:
            return render(
                request,
                "payments/db_error.html",
                {"error": exc, "host": settings.DATABASES["cba"]["HOST"]},
                status=503,
            )


class PaymentRequestListView(LoginRequiredMixin, DatabaseErrorMixin, FilterView):
    model = PaymentRequest
    filterset_class = PaymentRequestFilter
    template_name = "payments/list.html"
    context_object_name = "payment_requests"

    def get_paginate_by(self, queryset):
        try:
            size = int(self.request.GET.get("per_page", DEFAULT_PAGE_SIZE))
        except (TypeError, ValueError):
            return DEFAULT_PAGE_SIZE
        return size if size in PAGE_SIZES else DEFAULT_PAGE_SIZE

    def _query(self, **overrides):
        """Current query string with `overrides` applied and blanks dropped."""
        params = self.request.GET.copy()
        for key, value in overrides.items():
            if value is None:
                params.pop(key, None)
            else:
                params[key] = value
        for key in [k for k, v in params.items() if not v]:
            params.pop(key)
        return params.urlencode()

    def _sort_links(self):
        """Header links that toggle asc/desc, keeping filters intact."""
        current = self.request.GET.get("sort", "")
        links = {}
        for name, _label, sortable in LIST_COLUMNS:
            if not sortable:
                continue
            ascending = current == name
            links[name] = {
                "url": self._query(sort=f"-{name}" if ascending else name, page=None),
                "direction": "asc" if ascending else ("desc" if current == f"-{name}" else ""),
            }
        return links

    def _active_filters(self):
        """One chip per applied filter, each with a link that removes it."""
        chips = []
        for name, value in self.request.GET.items():
            if name in NON_FILTER_PARAMS or not value:
                continue
            field = self.filterset.form.fields.get(name)
            label = getattr(field, "label", None) or name.replace("_", " ").capitalize()
            # Choices may be a lazy/callable-backed object, so iterate it
            # rather than testing it for truthiness or length.
            display = value
            try:
                choices = {str(k): v for k, v in getattr(field, "choices", ())}
            except (TypeError, ValueError):
                choices = {}
            display = choices.get(value, value)
            chips.append({
                "label": label,
                "value": display,
                "remove_url": self._query(**{name: None, "page": None}),
            })
        return chips

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["list_columns"] = LIST_COLUMNS
        context["sort_links"] = self._sort_links()
        context["active_filters"] = self._active_filters()
        context["page_sizes"] = PAGE_SIZES
        context["per_page"] = self.get_paginate_by(None)
        context["page_size_links"] = {
            size: self._query(per_page=size, page=None) for size in PAGE_SIZES
        }
        context["summary"] = self.filterset.qs.aggregate(
            total=Count("id"),
            pending=Count("id", filter=Q(status="PENDING")),
            settled=Count("id", filter=Q(status__in=SETTLED_STATUSES)),
            total_value=Sum("request_amount"),
        )
        context["total_value_compact"] = compact_number(context["summary"]["total_value"])

        # Percentages for the tiles. Guarded against an empty result set.
        total = context["summary"]["total"] or 0
        context["pending_pct"] = round(context["summary"]["pending"] * 100 / total) if total else 0
        context["settled_pct"] = round(context["summary"]["settled"] * 100 / total) if total else 0

        page = context.get("page_obj")
        if page and total:
            context["range_start"] = page.start_index()
            context["range_end"] = page.end_index()
        params = self.request.GET.copy()
        params.pop("page", None)
        context["filter_query"] = params.urlencode()
        context["applied_filter_count"] = len(context["active_filters"])
        return context


class PaymentRequestDetailView(LoginRequiredMixin, DatabaseErrorMixin, DetailView):
    model = PaymentRequest
    template_name = "payments/detail.html"
    context_object_name = "payment_request"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        obj = self.object
        labels = {f.name: f.verbose_name for f in obj._meta.fields}

        humanised = {"request_type", "payment_type"}

        groups = []
        for title, names in DETAIL_GROUPS:
            rows = []
            for name in names:
                value = getattr(obj, name)
                if name in humanised:
                    value = PaymentRequest._humanise(value) or None
                elif isinstance(value, (dict, list)):
                    value = json.dumps(value, indent=2, ensure_ascii=False)
                rows.append((labels[name].capitalize(), value, name))
            groups.append((title, rows))
        context["groups"] = groups
        return context


def _cell(value, column):
    """Render one value for export."""
    if not settings.EXPORT_INCLUDE_PII and column in PII_COLUMNS and value:
        return "***"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    if value is None:
        return ""
    return value


def _filtered_queryset(request):
    filterset = PaymentRequestFilter(request.GET, queryset=PaymentRequest.objects.all())
    return filterset.qs


def _filename(extension):
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return f"payment-requests-{stamp}.{extension}"


@login_required
@handle_db_errors
def export_xlsx(request):
    """Export the current filter selection as a formatted Excel workbook.

    Unlike the CSV, this carries column widths and number formats, so dates and
    amounts are readable without the reader resizing anything.
    """
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Payment Requests")

    # Widths must be set before any row is appended in write-only mode.
    for index, name in enumerate(EXPORT_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = COLUMN_WIDTHS.get(
            name, DEFAULT_WIDTH
        )
    sheet.freeze_panes = "A2"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F1A60")
    header = []
    for name in EXPORT_COLUMNS:
        cell = WriteOnlyCell(sheet, value=name.replace("_", " ").capitalize())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        header.append(cell)
    sheet.append(header)

    rows = _filtered_queryset(request).values_list(*EXPORT_COLUMNS)
    _audit_export(request, "xlsx", rows.count())

    for row in rows.iterator(chunk_size=500):
        cells = []
        for value, name in zip(row, EXPORT_COLUMNS):
            value = _excel_cell(value, name)
            if name in DATE_COLUMNS or name in AMOUNT_COLUMNS:
                cell = WriteOnlyCell(sheet, value=value)
                cell.number_format = (
                    DATE_FORMAT if name in DATE_COLUMNS else AMOUNT_FORMAT
                )
                cells.append(cell)
            else:
                cells.append(value)
        sheet.append(cells)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{_filename("xlsx")}"'
    workbook.save(response)
    return response


def compact_number(value):
    """1284 -> 1,284 | 1284000 -> 1.3M. For stat tiles, not table columns."""
    if value is None:
        return "0"
    number = float(value)
    for limit, suffix in ((1e12, "T"), (1e9, "B"), (1e6, "M"), (1e3, "K")):
        if abs(number) >= limit:
            trimmed = f"{number / limit:.1f}".removesuffix(".0")
            return f"{trimmed}{suffix}"
    return f"{number:,.0f}"


def _audit_export(request, fmt, row_count):
    """Record who exported what.

    The exported columns include customer emails, phone numbers and account
    numbers, so downloads must be attributable after the fact.
    """
    filters = request.GET.urlencode() or "(none)"
    audit.info(
        "EXPORT format=%s user=%s ip=%s rows=%d pii=%s filters=%s",
        fmt,
        request.user.get_username(),
        request.META.get("REMOTE_ADDR", "?"),
        row_count,
        "included" if settings.EXPORT_INCLUDE_PII else "masked",
        filters,
    )


def _excel_cell(value, column):
    value = _cell(value, column)
    if hasattr(value, "quantize"):
        return float(value)
    return value
